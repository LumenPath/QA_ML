import pandas as pd
import numpy as np
import matplotlib

matplotlib.use('Agg')
import os
import time
import json
import warnings
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment
import joblib
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from sklearn.model_selection import train_test_split, KFold, cross_validate

# Import the visualization module (chart functions provided by the second file)
from visualization_module import (
    plot_academic_scatter, plot_enhanced_feature_importance,
    plot_side_by_side_scatter, export_feature_importance,
    export_predictions_to_csv
)

# Import our custom modules
from data_generator import ImprovedDataGenerator
from deterministic_ssl import DeterministicSelfTrainingRegressor

# Import the high-precision data storage module
from precision_storage import (
    save_data_with_high_precision,
    generate_visualizations_from_parquet,
    load_data_from_parquet,
    train_and_evaluate_from_parquet
)

# Ignore warnings
warnings.filterwarnings('ignore')

# ===========================
# Global parameter settings
# ===========================

# Equal-width sampling parameters
BINS = 5  # Number of equal-width sampling bins, initial value is 2

# Input file settings
INPUT_FILE = 'XXX.csv'

# Data column index configuration (soft-coded)
ID_COL_INDEX = 0  # ID column index
TARGET_COL_INDEX = 1  # Target variable column index
FEATURE_START_INDEX = 2  # Feature start column index
FEATURE_END_INDEX = 10  # Feature end column index, using 14 features (2-15 inclusive)

# Data splitting parameters
TEST_SIZE = 0.1  # Test set ratio
RANDOM_STATE = 42  # Random seed to ensure reproducible results
CV_FOLDS = 5  # Number of cross-validation folds

# Random forest parameters
RF_N_ESTIMATORS = 49  # Number of trees in the random forest
RF_N_JOBS = -1  # Number of CPU cores to use (-1 means using all available cores)

# Semi-supervised learning parameters
SSL_MAX_ITER = 50  # Maximum number of iterations
SSL_BATCH_SIZE = 10  # Number of samples added in each iteration
SSL_CONFIDENCE_THRESHOLD = 0.95  # Confidence threshold
SSL_USE_DISTANCE = True  # Whether to use distance to calculate confidence
SSL_N_NEIGHBORS = 3  # Number of nearest neighbors

# Confidence calculation weights
ENSEMBLE_WEIGHT = 0.42  # Weight of the ensemble learning method
DISTANCE_WEIGHT = 0.58  # Weight of the distance method
DENSITY_WEIGHT = 0.0  # Weight of the density method

# Data generation model parameters
GEN_DATA_SIZE = 20000  # Amount of generated data
GEN_RANDOM_STATE = 42  # Random seed to ensure reproducible generated results

# Output settings
BASE_OUTPUT_DIR = 'semi_supervised_learning_training_results'  # Base output directory
GENERATED_DATA_DIR = 'generated_data'  # Fixed folder name for generated data

# Column name settings
ID_COLUMN_NAME = 'materials'
TARGET_COLUMN_NAME = 'property'


# ===========================
# Data loading, cleaning, and splitting functions
# ===========================

def load_and_clean_data(file_path):
    """Load data and clean it"""
    # Load data
    df = pd.read_csv(file_path)
    print("Number of samples before data cleaning:", len(df))

    # Check and handle missing values
    if df.isnull().sum().any():
        print("Missing values found, processing...")
        df_cleaned = df.dropna()
        print("Number of samples after handling missing values:", len(df_cleaned))
    else:
        df_cleaned = df.copy()
        print("No missing values in the data")

    return df_cleaned


def equal_width_sampling(X, y, test_size=0.1, random_state=42, bins=BINS):
    """
    Equal-width sampling method - use equal-width binning for stratified sampling

    Parameters:
    - X: Feature data
    - y: Target variable
    - test_size: Test set ratio
    - random_state: Random seed
    - bins: Number of bins

    Returns:
    - X_train, X_test, y_train, y_test: Training set and test set
    """
    try:
        # Use the pandas pd.cut function to divide target variable y into the specified number of bins according to equal-width rules
        bin_edges = pd.cut(y, bins=bins, retbins=True)[1]
        y_bins = pd.cut(y, bins=bins, labels=False)

        # Use scikit-learn's train_test_split function for stratified sampling
        # Ensure samples in each bin are proportionally assigned to the training set and test set through the stratify parameter
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=test_size,
            stratify=y_bins, random_state=random_state
        )

        print(f"Equal-width sampling completed, number of bins used: {bins}")
        print(f"Bin edges: {bin_edges}")
        print(f"Training set size: {len(X_train)}, test set size: {len(X_test)}")

        # Print the sample distribution in each bin
        train_bins = pd.cut(y_train, bins=bin_edges, labels=False)
        test_bins = pd.cut(y_test, bins=bin_edges, labels=False)

        print("Number of samples in each bin in the training set:")
        train_bin_counts = pd.Series(train_bins).value_counts().sort_index()
        for bin_idx, count in train_bin_counts.items():
            print(f"  Bin {bin_idx}: {count} samples")

        print("Number of samples in each bin in the test set:")
        test_bin_counts = pd.Series(test_bins).value_counts().sort_index()
        for bin_idx, count in test_bin_counts.items():
            print(f"  Bin {bin_idx}: {count} samples")

        return X_train, X_test, y_train, y_test
    except Exception as e:
        print(f"Equal-width sampling failed: {str(e)}")
        traceback.print_exc()
        return None, None, None, None


def train_and_evaluate_model(X_train, X_test, y_train, y_test, n_folds=5, random_state=42, method_name="Model"):
    """
    Train a random forest model and evaluate it
    - Use n-fold cross-validation
    - Calculate R², MAE, and RMSE
    - Return the trained model and evaluation metrics
    """
    # Create the random forest model
    rf_model = RandomForestRegressor(n_estimators=RF_N_ESTIMATORS, random_state=random_state, n_jobs=RF_N_JOBS)

    print(f"\nStarting {n_folds}-fold cross-validation [{method_name}]...")

    # Use sklearn's built-in cross-validation function
    scoring = ['r2', 'neg_mean_absolute_error', 'neg_root_mean_squared_error']
    cv_results = cross_validate(
        rf_model, X_train, y_train,
        cv=n_folds,
        scoring=scoring,
        return_train_score=False,
        n_jobs=RF_N_JOBS
    )

    # Extract cross-validation results
    cv_r2_scores = cv_results['test_r2']
    cv_mae_scores = -cv_results['test_neg_mean_absolute_error']  # Convert to positive values
    cv_rmse_scores = -cv_results['test_neg_root_mean_squared_error']  # Convert to positive values

    # Print the results for each fold - do not limit decimal places
    for fold in range(n_folds):
        print(f"  Fold {fold + 1}: R² = {cv_r2_scores[fold]}, "
              f"MAE = {cv_mae_scores[fold]}, RMSE = {cv_rmse_scores[fold]}")

    # Output the average cross-validation results - do not limit decimal places
    print(f"\n{method_name} - Average cross-validation results:")
    print(f"Average validation R²: {np.mean(cv_r2_scores)} (±{np.std(cv_r2_scores)})")
    print(f"Average validation MAE: {np.mean(cv_mae_scores)} (±{np.std(cv_mae_scores)})")
    print(f"Average validation RMSE: {np.mean(cv_rmse_scores)} (±{np.std(cv_rmse_scores)})")

    # Train the final model on the full training set
    final_model = RandomForestRegressor(n_estimators=RF_N_ESTIMATORS, random_state=random_state, n_jobs=RF_N_JOBS)
    final_model.fit(X_train, y_train)

    # Calculate training set metrics
    y_train_pred = final_model.predict(X_train)
    train_r2 = r2_score(y_train, y_train_pred)
    train_mae = mean_absolute_error(y_train, y_train_pred)
    train_rmse = np.sqrt(mean_squared_error(y_train, y_train_pred))

    print(f"\n{method_name} - Training set evaluation results:")
    print(f"R²: {train_r2:.9f}")
    print(f"MAE: {train_mae:.9f}")
    print(f"RMSE: {train_rmse:.9f}")

    # Evaluate on the test set
    y_test_pred = final_model.predict(X_test)
    test_r2 = r2_score(y_test, y_test_pred)
    test_mae = mean_absolute_error(y_test, y_test_pred)
    test_rmse = np.sqrt(mean_squared_error(y_test, y_test_pred))

    print(f"\n{method_name} - Test set evaluation results:")
    print(f"R²: {test_r2:.9f}")
    print(f"MAE: {test_mae:.9f}")
    print(f"RMSE: {test_rmse:.9f}")

    # Create the results dictionary
    results = {
        'model': final_model,
        'metrics': {
            'train': {'r2': train_r2, 'mae': train_mae, 'rmse': train_rmse},
            'val': {
                'r2': np.mean(cv_r2_scores),
                'mae': np.mean(cv_mae_scores),
                'rmse': np.mean(cv_rmse_scores)
            },
            'test': {'r2': test_r2, 'mae': test_mae, 'rmse': test_rmse}
        },
        'pred': {
            'train': y_train_pred,
            'test': y_test_pred
        },
        'true': {
            'train': y_train,
            'test': y_test
        }
    }

    return results


# ===========================
# Helper functions
# ===========================

def setup_output_directories():
    """Create the output directory structure"""
    if not os.path.exists(BASE_OUTPUT_DIR):
        os.makedirs(BASE_OUTPUT_DIR)

    # Create the fixed generated data directory
    gen_data_dir = os.path.join(BASE_OUTPUT_DIR, GENERATED_DATA_DIR)
    if not os.path.exists(gen_data_dir):
        os.makedirs(gen_data_dir)

    # Create a subdirectory named with a timestamp
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    run_dir = os.path.join(BASE_OUTPUT_DIR, timestamp)
    os.makedirs(run_dir)

    # Create model and chart subdirectories
    models_dir = os.path.join(run_dir, "models")
    os.makedirs(models_dir)

    charts_dir = os.path.join(run_dir, "charts")
    os.makedirs(charts_dir)

    data_dir = os.path.join(run_dir, "data")
    os.makedirs(data_dir)

    return run_dir, models_dir, charts_dir, data_dir, gen_data_dir


def save_parameters(output_dir):
    """Save the parameter settings of the current run"""
    params = {
        "Number of equal-width sampling bins": BINS,
        "Input file": INPUT_FILE,
        "ID column index": ID_COL_INDEX,
        "Target variable column index": TARGET_COL_INDEX,
        "Feature start column index": FEATURE_START_INDEX,
        "Feature end column index": FEATURE_END_INDEX,
        "Test set ratio": TEST_SIZE,
        "Random seed": RANDOM_STATE,
        "Number of cross-validation folds": CV_FOLDS,
        "Number of random forest trees": RF_N_ESTIMATORS,
        "Number of CPU cores": RF_N_JOBS,
        "Maximum number of semi-supervised iterations": SSL_MAX_ITER,
        "Number of samples added per iteration": SSL_BATCH_SIZE,
        "Confidence threshold": SSL_CONFIDENCE_THRESHOLD,
        "Use distance to calculate confidence": SSL_USE_DISTANCE,
        "Number of nearest neighbors": SSL_N_NEIGHBORS,
        "Ensemble learning weight": ENSEMBLE_WEIGHT,
        "Distance method weight": DISTANCE_WEIGHT,
        "Density method weight": DENSITY_WEIGHT,
        "Generated data size": GEN_DATA_SIZE,
        "Generated data random seed": GEN_RANDOM_STATE
    }

    with open(os.path.join(output_dir, 'parameters.json'), 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False, indent=4)


def save_ssl_iterations_to_excel(output_dir, iteration_metrics, best_mae_iter):
    """Save the performance metrics of each semi-supervised learning iteration to Excel"""
    excel_path = os.path.join(output_dir, 'ssl_iterations.xlsx')

    # Create the data frame
    rows = []
    for iter_data in iteration_metrics:
        iteration = iter_data['iteration']
        metrics = iter_data['metrics']
        added_samples = iter_data['added_samples']
        total_pseudo_samples = iter_data['total_pseudo_samples']

        # Extract test set and training set metrics
        train_metrics = metrics.get('train', {})
        test_metrics = metrics.get('test', {})

        # Mark the best model iteration
        is_best_mae = iteration == best_mae_iter

        row = {
            'Iteration': iteration + 1,  # Start from 1
            'Added Samples': added_samples,
            'Total Pseudo Samples': total_pseudo_samples,
            'Train R²': train_metrics.get('r2', 'N/A'),
            'Train MAE': train_metrics.get('mae', 'N/A'),
            'Train RMSE': train_metrics.get('rmse', 'N/A'),
            'Test R²': test_metrics.get('r2', 'N/A'),
            'Test MAE': test_metrics.get('mae', 'N/A'),
            'Test RMSE': test_metrics.get('rmse', 'N/A'),
            'Best MAE Model': "YES" if is_best_mae else ""
        }
        rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Save to Excel
    df.to_excel(excel_path, index=False, sheet_name='SSL Iterations')

    # Beautify Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 15

    # Set header format
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Save the formatted Excel
    wb.save(excel_path)

    print(f"SSL iteration performance metrics have been saved to the Excel file: {excel_path}")
    return excel_path


def save_performance_metrics_excel(output_dir, base_metrics, ssl_metrics):
    """Save performance metrics to an Excel file with 9-decimal precision"""
    excel_path = os.path.join(output_dir, 'performance_metrics.xlsx')

    # Create the data frame
    metrics_data = {
        'Metric': ['Training R²', 'Training RMSE', 'Training MAE',
                   'Validation R²', 'Validation RMSE', 'Validation MAE',
                   'Testing R²', 'Testing RMSE', 'Testing MAE'],
        'Base Model': [
            base_metrics['train']['r2'],
            base_metrics['train']['rmse'],
            base_metrics['train']['mae'],
            base_metrics['val']['r2'],
            base_metrics['val']['rmse'],
            base_metrics['val']['mae'],
            base_metrics['test']['r2'],
            base_metrics['test']['rmse'],
            base_metrics['test']['mae']
        ],
        'SSL MAE Best Model': [
            ssl_metrics['r2_train'],
            ssl_metrics['rmse_train'],
            ssl_metrics['mae_train'],
            'N/A',  # The SSL model does not have validation set metrics
            'N/A',
            'N/A',
            ssl_metrics['r2_test'],
            ssl_metrics['rmse_test'],
            ssl_metrics['mae_test']
        ],
        'Improvement': [
            ssl_metrics['r2_train'] - base_metrics['train']['r2'],
            base_metrics['train']['rmse'] - ssl_metrics['rmse_train'],  # Smaller is better, so base model minus SSL model
            base_metrics['train']['mae'] - ssl_metrics['mae_train'],  # Smaller is better, so base model minus SSL model
            'N/A',
            'N/A',
            'N/A',
            ssl_metrics['r2_test'] - base_metrics['test']['r2'],
            base_metrics['test']['rmse'] - ssl_metrics['rmse_test'],  # Smaller is better, so base model minus SSL model
            base_metrics['test']['mae'] - ssl_metrics['mae_test']  # Smaller is better, so base model minus SSL model
        ]
    }

    # Create DataFrame
    df = pd.DataFrame(metrics_data)

    # Save to Excel - keep 9 decimal places
    df.to_excel(excel_path, index=False, sheet_name='Performance Metrics', float_format='%.9f')

    # Beautify the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Set header format
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Save the formatted Excel
    wb.save(excel_path)

    print(f"Performance metrics have been saved to the Excel file: {excel_path}")
    return excel_path


def check_existing_generated_data(gen_data_dir):
    """Check whether generated data files already exist"""
    if not os.path.exists(gen_data_dir):
        return None

    csv_files = [f for f in os.listdir(gen_data_dir) if f.endswith('.csv') and f.startswith('generated_data')]

    if not csv_files:
        return None

    # Sort by file modification time and select the latest file
    csv_files.sort(key=lambda x: os.path.getmtime(os.path.join(gen_data_dir, x)), reverse=True)
    latest_file = csv_files[0]

    return os.path.join(gen_data_dir, latest_file)


def load_generated_data(file_path, feature_names):
    """Load generated data"""
    try:
        df = pd.read_csv(file_path)

        # Extract target variable and features
        y_generated = df[TARGET_COLUMN_NAME].values

        # Ensure all required features exist
        missing_features = [f for f in feature_names if f not in df.columns]
        if missing_features:
            print(f"Warning: The following features are missing in the generated data: {missing_features}")
            return None, None

        X_generated = df[feature_names].values

        print(f"Generated data loaded successfully, shape: X={X_generated.shape}, y={y_generated.shape}")
        return X_generated, y_generated

    except Exception as e:
        print(f"Failed to load generated data: {e}")
        return None, None


def save_additional_csv_files(X_train, y_train, X_test, y_test, feature_names,
                              all_added_samples_X=None, all_added_samples_y=None,
                              best_model_samples_X=None, best_model_samples_y=None,
                              iteration_samples=None,
                              data_dir=None,
                              original_df=None,
                              id_col_index=0, target_col_index=1):
    """
    Function for saving the training set, test set, and pseudo-label data - modified version, saved with 9-decimal precision
    Only output the merged training set (original training set + best pseudo-labels) and do not separately generate prediction results for the original training set
    """
    import os
    import pandas as pd
    import numpy as np

    if data_dir is None:
        print("No data directory provided, skipping saving additional CSV files")
        return {}

    # Ensure the directory exists
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    # File names
    train_file = os.path.join(data_dir, 'original_train_set.csv')
    test_file = os.path.join(data_dir, 'original_test_set.csv')
    all_added_data_file = os.path.join(data_dir, 'all_pseudo_labeled_data.csv')
    best_model_data_file = os.path.join(data_dir, 'best_model_pseudo_labeled_data.csv')
    merged_file = os.path.join(data_dir, 'merged_train_and_best_model_data.csv')
    iterations_data_file = os.path.join(data_dir, 'iteration_added_samples.csv')

    # Added: Training set and test set files for the final semi-supervised learning model
    final_ssl_train_file = os.path.join(data_dir, 'final_ssl_training_set.csv')
    final_ssl_test_file = os.path.join(data_dir, 'final_ssl_test_set.csv')

    result_files = {
        'train_file': train_file,
        'test_file': test_file
    }

    # Get column names and ID name
    ID_COLUMN_NAME = 'materials' if original_df is None else original_df.columns[id_col_index]
    TARGET_COLUMN_NAME = 'property' if original_df is None else original_df.columns[target_col_index]

    print(f"Using ID column name: {ID_COLUMN_NAME}")
    print(f"Using target column name: {TARGET_COLUMN_NAME}")
    print(f"Using feature column names: {feature_names}")

    # Define a function to match samples with the original data to obtain the original ID
    def match_sample_with_original(sample, original_df, feature_names, tolerance=1e-8):
        """Match a sample with the original data and return the matching index or None"""
        if original_df is None:
            return None

        for idx, row in original_df.iterrows():
            match = True
            for feat in feature_names:
                # Get the sample feature value
                if isinstance(sample, pd.Series) or isinstance(sample, pd.DataFrame):
                    sample_value = sample[feat]
                elif isinstance(sample, np.ndarray):
                    feat_idx = feature_names.index(feat)
                    sample_value = sample[feat_idx]
                else:
                    match = False
                    break

                # Get the feature value from the original row
                orig_value = row[feat]

                # Compare values while considering floating-point errors
                if abs(float(sample_value) - float(orig_value)) > tolerance:
                    match = False
                    break

            if match:
                return idx
        return None

    # 1. Save the training set - use 9-decimal precision
    train_df = pd.DataFrame()

    # Try to get the original IDs if the original dataset is available
    train_ids = []
    if original_df is not None:
        if isinstance(X_train, pd.DataFrame):
            print("Matching training set samples using the DataFrame method...")
            for idx, row in X_train.iterrows():
                orig_idx = match_sample_with_original(row, original_df, feature_names)
                if orig_idx is not None:
                    train_ids.append(original_df.iloc[orig_idx][ID_COLUMN_NAME])
                else:
                    train_ids.append(f"train_{idx}")
        else:
            print("Matching training set samples using the NumPy array method...")
            for i, sample in enumerate(X_train):
                orig_idx = match_sample_with_original(sample, original_df, feature_names)
                if orig_idx is not None:
                    train_ids.append(original_df.iloc[orig_idx][ID_COLUMN_NAME])
                else:
                    train_ids.append(f"train_{i}")
    else:
        # If there is no original data, use generated IDs
        train_ids = [f"train_{i}" for i in range(len(y_train))]

    # Add the ID column
    train_df[ID_COLUMN_NAME] = train_ids

    # Add the target variable column - use 9-decimal precision
    if isinstance(y_train, pd.Series):
        train_df[TARGET_COLUMN_NAME] = y_train.values
    else:
        train_df[TARGET_COLUMN_NAME] = y_train

    # Add feature columns - use 9-decimal precision
    if isinstance(X_train, pd.DataFrame):
        for feature in feature_names:
            train_df[feature] = X_train[feature].values
    else:
        for i, feature in enumerate(feature_names):
            train_df[feature] = X_train[:, i]

    # Verify data integrity
    print(f"Training set shape: {train_df.shape}")
    print(f"Number of missing values in the training set: {train_df.isnull().sum().sum()}")

    # Save - use 9-decimal precision
    train_df.to_csv(train_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
    print(f"Original training set has been saved to: {train_file}")

    # 2. Save the test set - use 9-decimal precision
    test_df = pd.DataFrame()

    # Try to get the original IDs if the original dataset is available
    test_ids = []
    if original_df is not None:
        if isinstance(X_test, pd.DataFrame):
            print("Matching test set samples using the DataFrame method...")
            for idx, row in X_test.iterrows():
                orig_idx = match_sample_with_original(row, original_df, feature_names)
                if orig_idx is not None:
                    test_ids.append(original_df.iloc[orig_idx][ID_COLUMN_NAME])
                else:
                    test_ids.append(f"test_{idx}")
        else:
            print("Matching test set samples using the NumPy array method...")
            for i, sample in enumerate(X_test):
                orig_idx = match_sample_with_original(sample, original_df, feature_names)
                if orig_idx is not None:
                    test_ids.append(original_df.iloc[orig_idx][ID_COLUMN_NAME])
                else:
                    test_ids.append(f"test_{i}")
    else:
        # If there is no original data, use generated IDs
        test_ids = [f"test_{i}" for i in range(len(y_test))]

    # Add the ID column
    test_df[ID_COLUMN_NAME] = test_ids

    # Add the target variable column - use 9-decimal precision
    if isinstance(y_test, pd.Series):
        test_df[TARGET_COLUMN_NAME] = y_test.values
    else:
        test_df[TARGET_COLUMN_NAME] = y_test

    # Check whether the target variable is complete
    if pd.isna(test_df[TARGET_COLUMN_NAME]).any():
        print("Warning: The test set target variable contains NaN values and will be filled with -999")
        test_df[TARGET_COLUMN_NAME] = test_df[TARGET_COLUMN_NAME].fillna(-999)

    # Add feature columns - use 9-decimal precision
    if isinstance(X_test, pd.DataFrame):
        for feature in feature_names:
            test_df[feature] = X_test[feature].values
    else:
        for i, feature in enumerate(feature_names):
            test_df[feature] = X_test[:, i]

    # Verify data integrity
    print(f"Test set shape: {test_df.shape}")
    print(f"Number of target variable values in the test set: {len(test_df[TARGET_COLUMN_NAME])}")
    print(f"Number of missing values in the test set: {test_df.isnull().sum().sum()}")

    # Save - use 9-decimal precision
    test_df.to_csv(test_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
    print(f"Original test set has been saved to: {test_file}")

    # 3. Save all pseudo-label data - if any
    if all_added_samples_X is not None and all_added_samples_y is not None and len(all_added_samples_X) > 0:
        pseudo_df = pd.DataFrame()

        # Add the ID column
        pseudo_df[ID_COLUMN_NAME] = [f"pseudo_{i}" for i in range(len(all_added_samples_y))]

        # Add the target variable column - use 9-decimal precision
        if isinstance(all_added_samples_y, (list, tuple)):
            pseudo_df[TARGET_COLUMN_NAME] = all_added_samples_y
        else:
            pseudo_df[TARGET_COLUMN_NAME] = list(all_added_samples_y)

        # Add feature columns - use 9-decimal precision
        for i, feature in enumerate(feature_names):
            feature_values = []

            if isinstance(all_added_samples_X[0], np.ndarray):
                feature_values = [x[i] for x in all_added_samples_X]
            elif isinstance(all_added_samples_X[0], pd.Series):
                feature_values = [x[feature] if feature in x else x.iloc[i] for x in all_added_samples_X]
            elif isinstance(all_added_samples_X[0], pd.DataFrame):
                feature_values = [x[feature].iloc[0] if feature in x else 0 for x in all_added_samples_X]
            else:
                try:
                    feature_values = [x[i] if isinstance(x, (list, tuple, np.ndarray)) else
                                      getattr(x, feature, 0) for x in all_added_samples_X]
                except:
                    print(f"Warning: Unable to extract feature {feature} from pseudo-label data, filling with 0")
                    feature_values = [0] * len(all_added_samples_y)

            # Add feature values
            pseudo_df[feature] = feature_values

        # Save - use 9-decimal precision
        pseudo_df.to_csv(all_added_data_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
        print(f"All pseudo-label data has been saved to: {all_added_data_file}")
        result_files['all_added_data_file'] = all_added_data_file

    # 4. Save the pseudo-label data used by the best model - if any
    if best_model_samples_X is not None and best_model_samples_y is not None and len(best_model_samples_X) > 0:
        best_df = pd.DataFrame()

        # Add the ID column
        best_df[ID_COLUMN_NAME] = [f"best_pseudo_{i}" for i in range(len(best_model_samples_y))]

        # Add the target variable column - use 9-decimal precision
        if isinstance(best_model_samples_y, (list, tuple)):
            best_df[TARGET_COLUMN_NAME] = best_model_samples_y
        else:
            best_df[TARGET_COLUMN_NAME] = list(best_model_samples_y)

        # Add feature columns - use 9-decimal precision
        for i, feature in enumerate(feature_names):
            feature_values = []

            if isinstance(best_model_samples_X[0], np.ndarray):
                feature_values = [x[i] for x in best_model_samples_X]
            elif isinstance(best_model_samples_X[0], pd.Series):
                feature_values = [x[feature] if feature in x else x.iloc[i] for x in best_model_samples_X]
            elif isinstance(best_model_samples_X[0], pd.DataFrame):
                feature_values = [x[feature].iloc[0] if feature in x else 0 for x in best_model_samples_X]
            else:
                try:
                    feature_values = [x[i] if isinstance(x, (list, tuple, np.ndarray)) else
                                      getattr(x, feature, 0) for x in best_model_samples_X]
                except:
                    print(f"Warning: Unable to extract feature {feature} from the best model pseudo-label data, filling with 0")
                    feature_values = [0] * len(best_model_samples_y)

            # Add feature values
            best_df[feature] = feature_values

        # Save - use 9-decimal precision
        best_df.to_csv(best_model_data_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
        print(f"Best model pseudo-label data has been saved to: {best_model_data_file}")
        result_files['best_model_data_file'] = best_model_data_file

        # 5. Create merged data
        print("Creating merged data...")
        merged_df = pd.concat([train_df, best_df], ignore_index=True)

        # Verify the merging result
        print(f"Merged data shape: {merged_df.shape}")
        print(f"Expected number of rows: {len(train_df) + len(best_df)}")

        # Save merged data - use 9-decimal precision
        merged_df.to_csv(merged_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
        print(f"Merged data has been saved to: {merged_file}")
        result_files['merged_file'] = merged_file

        # Export the merged data as the training set of the final semi-supervised learning model - use 9-decimal precision
        print("Exporting the training set of the final semi-supervised learning model...")
        merged_df.to_csv(final_ssl_train_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
        print(f"The training set of the final semi-supervised learning model has been saved to: {final_ssl_train_file}")
        result_files['final_ssl_train_file'] = final_ssl_train_file

    # Export the test set of the final semi-supervised learning model (same as the original test set, but ensuring consistent format) - use 9-decimal precision
    print("Exporting the test set of the final semi-supervised learning model...")
    test_df.to_csv(final_ssl_test_file, index=False, float_format='%.9f', na_rep='NaN', lineterminator='\n')
    print(f"The test set of the final semi-supervised learning model has been saved to: {final_ssl_test_file}")
    result_files['final_ssl_test_file'] = final_ssl_test_file

    # 6. Save iteration sample data - if any
    if iteration_samples is not None and len(iteration_samples) > 0:
        # Create an empty DataFrame
        iterations_df = pd.DataFrame()
        rows = []

        # Collect samples from each iteration
        for iter_idx, iter_data in enumerate(iteration_samples):
            if 'added_X' in iter_data and 'added_y' in iter_data:
                added_X = iter_data['added_X']
                added_y = iter_data['added_y']

                if len(added_X) > 0:
                    for sample_idx, (x_sample, y_sample) in enumerate(zip(added_X, added_y)):
                        row = {
                            'Iteration': iter_idx + 1,
                            'Sample_ID': f"iter_{iter_idx + 1}_sample_{sample_idx + 1}"
                        }

                        # Add the target variable - use 9-decimal precision
                        row[TARGET_COLUMN_NAME] = y_sample

                        # Add features - use 9-decimal precision
                        for i, feature in enumerate(feature_names):
                            if isinstance(x_sample, np.ndarray):
                                feature_value = x_sample[i]
                            elif isinstance(x_sample, pd.Series):
                                feature_value = x_sample[feature] if feature in x_sample else x_sample.iloc[i]
                            elif isinstance(x_sample, pd.DataFrame):
                                feature_value = x_sample[feature].iloc[0] if feature in x_sample else 0
                            else:
                                try:
                                    feature_value = x_sample[i] if isinstance(x_sample, (list, tuple)) else getattr(
                                        x_sample, feature, 0)
                                except:
                                    feature_value = 0

                            row[feature] = feature_value

                        rows.append(row)

        # If rows have been collected, create a DataFrame and save it
        if rows:
            iterations_df = pd.DataFrame(rows)
            iterations_df.to_csv(iterations_data_file, index=False, float_format='%.9f', na_rep='NaN',
                                 lineterminator='\n')
            print(f"Iteration sample data has been saved to: {iterations_data_file}")
            result_files['iterations_data_file'] = iterations_data_file

    return result_files


def run_with_generated_data(base_results, X_train, y_train, X_test, y_test, feature_names, output_dirs,
                            original_df=None):
    """Execute semi-supervised learning with generated data - modified so that the best MAE model is no longer retrained and the original training set is no longer evaluated"""
    run_dir, models_dir, charts_dir, data_dir, gen_data_dir = output_dirs

    # Check whether generated data already exists
    existing_data_file = check_existing_generated_data(gen_data_dir)
    generate_new_data = True

    if existing_data_file:
        print(f"\nExisting generated data found: {os.path.basename(existing_data_file)}")
        choice = input("Use this data for semi-supervised learning? (y/n): ").lower().strip()

        if choice == 'y':
            generate_new_data = False
            X_generated, y_generated = load_generated_data(existing_data_file, feature_names)

            if X_generated is None:
                print("Unable to use existing data, new data will be generated...")
                generate_new_data = True

    if generate_new_data:
        print("\nGenerating new data and executing semi-supervised learning...")

        # Extract the base model
        base_model = base_results['model']

        # Create the training data DataFrame (for data generation)
        if isinstance(X_train, pd.DataFrame):
            train_df = X_train.copy()
            # Add the ID column and target variable column
            train_df.insert(ID_COL_INDEX, ID_COLUMN_NAME, [f"sample{i}" for i in range(len(y_train))])
            train_df.insert(TARGET_COL_INDEX, TARGET_COLUMN_NAME, y_train)
        else:
            # If X is a numpy array, create a new DataFrame
            train_df = pd.DataFrame()
            train_df[ID_COLUMN_NAME] = [f"sample{i}" for i in range(len(y_train))]  # Create the ID column
            train_df[TARGET_COLUMN_NAME] = y_train  # Target variable column

            # Add feature columns
            for i, feature_name in enumerate(feature_names):
                train_df[feature_name] = X_train[:, i]

        # Create the improved data generator
        print("Creating the improved data generator...")
        data_generator = ImprovedDataGenerator(
            input_data=train_df,
            output_dir=gen_data_dir,  # Use the fixed generated data directory
            n_generate=GEN_DATA_SIZE,
            random_state=GEN_RANDOM_STATE,
            id_col_index=ID_COL_INDEX,
            target_col_index=TARGET_COL_INDEX,
            feature_start_index=FEATURE_START_INDEX,
            feature_end_index=FEATURE_END_INDEX,
            id_column_name=ID_COLUMN_NAME,
            target_column_name=TARGET_COLUMN_NAME
        )

        # Generate data
        print("Starting the data generation process...")
        data_generator.load_data()
        data_generator.build_models()
        data_generator.generate_samples()
        data_generator.validate_generation()
        generated_file, generated_df = data_generator.save_results()

        # Get generated data
        X_generated, y_generated = data_generator.get_generated_data()
    else:
        generated_file = existing_data_file

    # If the training data is a DataFrame, also convert the generated data to a DataFrame
    if isinstance(X_train, pd.DataFrame):
        X_generated_df = pd.DataFrame(X_generated, columns=X_train.columns)
        X_generated = X_generated_df

    # Use the generated data for semi-supervised learning
    print("\nUsing generated data for semi-supervised learning...")

    # Semi-supervised model - use the deterministic version
    ssl_model = DeterministicSelfTrainingRegressor(
        base_estimator=RandomForestRegressor(
            n_estimators=RF_N_ESTIMATORS,
            random_state=RANDOM_STATE,
            n_jobs=1  # Important: set to 1 to ensure sequential execution
        ),
        max_iter=SSL_MAX_ITER,
        batch_size=SSL_BATCH_SIZE,
        confidence_threshold=SSL_CONFIDENCE_THRESHOLD,
        verbose=True,
        use_distance=SSL_USE_DISTANCE,
        n_neighbors=SSL_N_NEIGHBORS,
        ensemble_weight=ENSEMBLE_WEIGHT,
        distance_weight=DISTANCE_WEIGHT,
        density_weight=DENSITY_WEIGHT,
        X_test=X_test,  # Provide the test set for evaluation
        y_test=y_test,
        random_state=RANDOM_STATE  # Add the random seed parameter
    )

    # Train the semi-supervised model
    try:
        ssl_model.fit(X_train, y_train, X_generated)
    except Exception as e:
        print(f"An error occurred during semi-supervised learning training: {e}")
        traceback.print_exc()
        return None, generated_file, None, None, None

    # Save iteration metrics to Excel
    iterations_excel = save_ssl_iterations_to_excel(
        run_dir,
        ssl_model.iteration_metrics,
        ssl_model.best_mae_iter
    )

    # Get the best MAE model - no longer retrain, directly use the saved best model
    print("\nGetting and evaluating the best MAE model...")
    try:
        best_mae_model = ssl_model.get_best_mae_model()
    except Exception as e:
        print(f"An error occurred while getting the best MAE model: {e}")
        traceback.print_exc()
        best_mae_model = base_results['model']  # If an error occurs, fall back to the base model

    # Directly evaluate the performance of the best model on the test set and merged training set
    try:
        # Prepare the merged training set (original training set + pseudo-label samples from the best model)
        best_pseudo_X = ssl_model.best_model_samples_X
        best_pseudo_y = ssl_model.best_model_samples_y

        # Evaluate on the test set
        y_test_pred_best = best_mae_model.predict(X_test)
        best_test_r2 = r2_score(y_test, y_test_pred_best)
        best_test_mae = mean_absolute_error(y_test, y_test_pred_best)
        best_test_rmse = np.sqrt(mean_squared_error(y_test, y_test_pred_best))

        if len(best_pseudo_X) > 0:
            # If there are pseudo-label samples, build the merged training set (same as during the iteration process)
            if isinstance(X_train, pd.DataFrame):
                # Merging for DataFrame type
                pseudo_df = pd.DataFrame(columns=X_train.columns)
                for i, sample in enumerate(best_pseudo_X):
                    if isinstance(sample, pd.Series):
                        pseudo_df = pd.concat([pseudo_df, pd.DataFrame([sample])], ignore_index=True)
                    elif isinstance(sample, np.ndarray):
                        pseudo_df.loc[i] = sample

                X_combined = pd.concat([X_train, pseudo_df], ignore_index=True)

                if isinstance(y_train, pd.Series):
                    y_combined = pd.concat([y_train, pd.Series(best_pseudo_y)], ignore_index=True)
                else:
                    y_combined = np.concatenate([y_train, best_pseudo_y])
            else:
                # Merging for numpy array type
                pseudo_X = np.array([np.array(x) for x in best_pseudo_X])
                X_combined = np.vstack([X_train, pseudo_X])
                y_combined = np.concatenate([y_train, best_pseudo_y])

            # Evaluate on the merged training set
            y_combined_pred_best = best_mae_model.predict(X_combined)
            best_train_r2 = r2_score(y_combined, y_combined_pred_best)
            best_train_mae = mean_absolute_error(y_combined, y_combined_pred_best)
            best_train_rmse = np.sqrt(mean_squared_error(y_combined, y_combined_pred_best))

            print(f"Best MAE model (from iteration {ssl_model.best_mae_iter + 1}) evaluation results:")
            print(
                f"  Merged training set (original + pseudo-label) R²: {best_train_r2:.9f}, MAE: {best_train_mae:.9f}, RMSE: {best_train_rmse:.9f}")
            print(f"  Test set R²: {best_test_r2:.9f}, MAE: {best_test_mae:.9f}, RMSE: {best_test_rmse:.9f}")
            print(f"  Number of pseudo-label samples used: {len(best_pseudo_X)}")
        else:
            # If there are no pseudo-label samples, directly use the original training set
            y_train_pred_best = best_mae_model.predict(X_train)
            best_train_r2 = r2_score(y_train, y_train_pred_best)
            best_train_mae = mean_absolute_error(y_train, y_train_pred_best)
            best_train_rmse = np.sqrt(mean_squared_error(y_train, y_train_pred_best))

            print(f"Best MAE model (from iteration {ssl_model.best_mae_iter + 1}) evaluation results:")
            print(f"  Training set R²: {best_train_r2:.9f}, MAE: {best_train_mae:.9f}, RMSE: {best_train_rmse:.9f}")
            print(f"  Test set R²: {best_test_r2:.9f}, MAE: {best_test_mae:.9f}, RMSE: {best_test_rmse:.9f}")
            print(f"  Number of pseudo-label samples used: 0")
    except Exception as e:
        print(f"An error occurred while evaluating the best MAE model: {e}")
        traceback.print_exc()
        # Use the metrics of the base model as backup
        best_train_r2 = base_results['metrics']['train']['r2']
        best_train_mae = base_results['metrics']['train']['mae']
        best_train_rmse = base_results['metrics']['train']['rmse']
        best_test_r2 = base_results['metrics']['test']['r2']
        best_test_mae = base_results['metrics']['test']['mae']
        best_test_rmse = base_results['metrics']['test']['rmse']

    # Create performance metrics for the best model
    ssl_mae_metrics = {
        'r2_train': best_train_r2,
        'rmse_train': best_train_rmse,
        'mae_train': best_train_mae,
        'r2_test': best_test_r2,
        'rmse_test': best_test_rmse,
        'mae_test': best_test_mae
    }

    # Use the visualization module to plot results
    try:
        # Use the academic-style function from the visualization module
        plot_enhanced_feature_importance(
            best_mae_model, feature_names, "Best MAE SSL Model - Feature Importance",
            os.path.join(charts_dir, "best_mae_feature_importance.png")
        )

        # Plot scatter plots for the training set and test set - only use the merged training set
        if len(best_pseudo_X) > 0:
            plot_academic_scatter(
                y_combined, y_combined_pred_best, "Best MAE SSL Model - Combined Training Set", best_train_r2,
                os.path.join(charts_dir, "best_mae_train_scatter.png")
            )
        else:
            plot_academic_scatter(
                y_train, y_train_pred_best, "Best MAE SSL Model - Training Set", best_train_r2,
                os.path.join(charts_dir, "best_mae_train_scatter.png")
            )

        plot_academic_scatter(
            y_test, y_test_pred_best, "Best MAE SSL Model - Test Set", best_test_r2,
            os.path.join(charts_dir, "best_mae_test_scatter.png")
        )

        # Plot side-by-side comparison of the training set and test set - only use the merged training set
        if len(best_pseudo_X) > 0:
            plot_side_by_side_scatter(
                y_combined, y_combined_pred_best, y_test, y_test_pred_best,
                "Best MAE SSL Model Evaluation (Combined Training Set)", best_train_r2, best_test_r2,
                os.path.join(charts_dir, "best_mae_side_by_side.png")
            )
        else:
            plot_side_by_side_scatter(
                y_train, y_train_pred_best, y_test, y_test_pred_best,
                "Best MAE SSL Model Evaluation", best_train_r2, best_test_r2,
                os.path.join(charts_dir, "best_mae_side_by_side.png")
            )

        # Export prediction results to CSV - only export merged training set and test set
        if len(best_pseudo_X) > 0:
            export_predictions_to_csv(
                y_combined, y_combined_pred_best, "best_mae_combined_training", charts_dir
            )
        else:
            export_predictions_to_csv(
                y_train, y_train_pred_best, "best_mae_training", charts_dir
            )

        export_predictions_to_csv(
            y_test, y_test_pred_best, "best_mae_test", charts_dir
        )

        # Export feature importance
        export_feature_importance(
            best_mae_model, feature_names, charts_dir
        )
    except Exception as e:
        print(f"An error occurred while plotting charts for the best MAE model: {e}")
        traceback.print_exc()

    # Compare the performance of the base model and the semi-supervised model
    base_metrics = base_results['metrics']

    # Save model comparison results - 9-decimal precision
    model_comparison = pd.DataFrame({
        'Model': ['Base Model', 'SSL Model (Best MAE)'],
        'Train_R2': [base_metrics['train']['r2'], ssl_mae_metrics['r2_train']],
        'Train_RMSE': [base_metrics['train']['rmse'], ssl_mae_metrics['rmse_train']],
        'Train_MAE': [base_metrics['train']['mae'], ssl_mae_metrics['mae_train']],
        'Test_R2': [base_metrics['test']['r2'], ssl_mae_metrics['r2_test']],
        'Test_RMSE': [base_metrics['test']['rmse'], ssl_mae_metrics['rmse_test']],
        'Test_MAE': [base_metrics['test']['mae'], ssl_mae_metrics['mae_test']]
    })

    model_comparison.to_csv(os.path.join(run_dir, 'model_comparison.csv'), index=False, float_format='%.9f')

    # Create an Excel performance metrics report (using the metrics of the MAE-optimal model)
    excel_path = save_performance_metrics_excel(run_dir, base_metrics, ssl_mae_metrics)

    print("\nModel comparison:")
    print(f"Base Model Test MAE: {base_metrics['test']['mae']:.9f}")
    print(f"Best MAE SSL Model Test MAE: {ssl_mae_metrics['mae_test']:.9f} (from iteration {ssl_model.best_mae_iter + 1})")
    print(f"MAE Improvement: {(base_metrics['test']['mae'] - ssl_mae_metrics['mae_test']):.9f}")

    # Add additional R2 metric comparison
    print(f"Base Model Test R²: {base_metrics['test']['r2']:.9f}")
    print(f"Best MAE SSL Model Test R²: {ssl_mae_metrics['r2_test']:.9f}")
    print(f"R² Improvement: {(ssl_mae_metrics['r2_test'] - base_metrics['test']['r2']):.9f}")

    # Save the MAE-best model to models_dir
    try:
        import joblib
        mae_model_path = os.path.join(models_dir, 'best_mae_model.joblib')
        joblib.dump(best_mae_model, mae_model_path)
        print(f"\nModel has been saved:")
        print(f"MAE-best model: {mae_model_path}")
    except Exception as e:
        print(f"Failed to save model: {e}")

    # Add high-precision data storage
    print("\nSaving high-precision model data...")
    try:
        # Directly use the best_model_samples_X and best_model_samples_y attributes of the SSL model
        # These are the pseudo-label samples actually used by the MAE-best model
        if ssl_model is not None:
            print(f"Using the training data of the MAE-best model (iteration {ssl_model.best_mae_iter + 1})")
            best_samples_X = ssl_model.best_model_samples_X
            best_samples_y = ssl_model.best_model_samples_y
            print(f"Using pseudo-label samples of the MAE-best model: {len(best_samples_X)}")
        else:
            best_samples_X = None
            best_samples_y = None
            print("No SSL model provided, unable to obtain pseudo-label samples of the MAE-best model")

        high_precision_files = save_data_with_high_precision(
            X_train=X_train,
            y_train=y_train,
            X_test=X_test,
            y_test=y_test,
            feature_names=feature_names,
            best_model=best_mae_model,
            all_pseudo_X=ssl_model.added_samples_X_,
            all_pseudo_y=ssl_model.added_samples_y_,
            best_pseudo_X=best_samples_X,
            best_pseudo_y=best_samples_y,
            ssl_model=ssl_model,
            data_dir=data_dir
        )

        # Generate visualization charts from Parquet files
        vis_files = generate_visualizations_from_parquet(
            high_precision_files['parquet_dir'],
            high_precision_files['model_file'],
            charts_dir,
            feature_names
        )

        print("\nHigh-precision data storage completed and can be used for subsequent analysis and training")

        # Add high-precision file paths to the results
        if not isinstance(high_precision_files, dict):
            high_precision_files = {}

        result_files = high_precision_files

    except Exception as e:
        print(f"An error occurred while saving high-precision data: {e}")
        traceback.print_exc()
        result_files = {}

    # Save additional CSV files
    try:
        # Directly use best_model_samples_X and best_model_samples_y of the semi-supervised model
        # These are the pseudo-label samples actually used by the MAE-best model
        if ssl_model is not None:
            best_model_samples_X = ssl_model.best_model_samples_X
            best_model_samples_y = ssl_model.best_model_samples_y
            print(f"\nUsing pseudo-label samples of the MAE-best model for CSV files: {len(best_model_samples_X)}")
        else:
            best_model_samples_X = None
            best_model_samples_y = None
            print("\nNo SSL model provided, CSV files will not contain pseudo-label samples")

        csv_files = save_additional_csv_files(
            X_train, y_train, X_test, y_test, feature_names,
            all_added_samples_X=ssl_model.added_samples_X_ if ssl_model is not None else None,
            all_added_samples_y=ssl_model.added_samples_y_ if ssl_model is not None else None,
            best_model_samples_X=best_model_samples_X,
            best_model_samples_y=best_model_samples_y,
            iteration_samples=ssl_model.iteration_samples if ssl_model is not None else None,
            data_dir=data_dir,
            original_df=original_df,
            id_col_index=ID_COL_INDEX,
            target_col_index=TARGET_COL_INDEX
        )
        print("\nAdditional CSV files saved successfully")

        # Merge high-precision file paths and CSV file paths
        for k, v in csv_files.items():
            result_files[k] = v

        # Specifically emphasize the newly added final semi-supervised learning data files
        if 'final_ssl_train_file' in csv_files and csv_files['final_ssl_train_file']:
            print("\nFinal model training and test data for semi-supervised learning:")
            print(f"Training set: {os.path.basename(csv_files['final_ssl_train_file'])}")
        if 'final_ssl_test_file' in csv_files and csv_files['final_ssl_test_file']:
            print(f"Test set: {os.path.basename(csv_files['final_ssl_test_file'])}")
        print("\nThese files maintain the same format and precision as other semi-supervised learning data.")

        # Specifically emphasize high-precision data files
        if 'hdf_file' in result_files and result_files['hdf_file']:
            print("\nHigh-precision data file (fully preserving training precision):")
            print(f"HDF5 file: {os.path.basename(result_files['hdf_file'])}")
            print("This HDF5 file contains all necessary data and accurately preserves the numerical precision during training.")

    except Exception as e:
        print(f"An error occurred while saving additional CSV files: {e}")
        traceback.print_exc()

    return best_mae_model, generated_file, excel_path, iterations_excel, result_files


# ===========================
# Main function
# ===========================

def main():
    """Main function - modified version, adding the function of reading final metrics from Parquet files and saving them"""
    print("Starting the improved data generation and semi-supervised learning integration framework...")

    # Create the output directory structure
    output_dirs = setup_output_directories()
    run_dir, models_dir, charts_dir, data_dir, gen_data_dir = output_dirs

    # Save parameter settings
    save_parameters(run_dir)

    # Load original data
    print("\n1. Loading original data...")
    df_cleaned = load_and_clean_data(INPUT_FILE)

    # Save original column names for later restoration
    original_columns = df_cleaned.columns.tolist()

    # Modify column names to standardized names
    column_mapping = {}
    if ID_COL_INDEX < len(original_columns):
        column_mapping[original_columns[ID_COL_INDEX]] = ID_COLUMN_NAME
    if TARGET_COL_INDEX < len(original_columns):
        column_mapping[original_columns[TARGET_COL_INDEX]] = TARGET_COLUMN_NAME

    # Apply column name mapping
    df_cleaned = df_cleaned.rename(columns=column_mapping)

    # Update feature column names
    updated_columns = df_cleaned.columns.tolist()

    # Prepare data - obtain features and target variable from the specified indexes
    target_column = updated_columns[TARGET_COL_INDEX]  # Target variable column name
    feature_columns = updated_columns[FEATURE_START_INDEX:FEATURE_END_INDEX + 1]  # List of feature column names

    print(f"Target variable: {target_column}")
    print(f"Feature variables: {', '.join(feature_columns[:5])}... ({len(feature_columns)} features in total)")

    X = df_cleaned[feature_columns]  # Use column names to select features
    y = df_cleaned[target_column]  # Use column name to select the target variable

    # Split the training set and test set
    print("\n2. Splitting the training set and test set...")
    X_train, X_test, y_train, y_test = equal_width_sampling(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, bins=BINS
    )
    print(f"Training set size: {X_train.shape[0]}")
    print(f"Test set size: {X_test.shape[0]}")

    # Create the analysis result saving file
    result_file = os.path.join(run_dir, "performance_metrics.csv")
    metrics_df = pd.DataFrame(columns=["Model", "Training_R2", "Training_MAE", "Training_RMSE",
                                       "Validation_R2", "Validation_MAE", "Validation_RMSE",
                                       "Testing_R2", "Testing_MAE", "Testing_RMSE"])

    # Train and evaluate the base model
    print("\n3. Evaluating the base model...")
    base_results = train_and_evaluate_model(
        X_train, X_test, y_train, y_test,
        n_folds=CV_FOLDS,
        random_state=RANDOM_STATE,
        method_name="Base Random Forest"
    )

    # Save base model analysis results - 9-decimal precision
    metrics_df = pd.concat([metrics_df, pd.DataFrame([{
        "Model": "RandomForest",
        "Training_R2": base_results['metrics']['train']['r2'],
        "Training_MAE": base_results['metrics']['train']['mae'],
        "Training_RMSE": base_results['metrics']['train']['rmse'],
        "Validation_R2": base_results['metrics']['val']['r2'],
        "Validation_MAE": base_results['metrics']['val']['mae'],
        "Validation_RMSE": base_results['metrics']['val']['rmse'],
        "Testing_R2": base_results['metrics']['test']['r2'],
        "Testing_MAE": base_results['metrics']['test']['mae'],
        "Testing_RMSE": base_results['metrics']['test']['rmse']
    }])], ignore_index=True)

    # Use the visualization module to plot various charts for the base model
    # Plot feature importance
    plot_enhanced_feature_importance(
        base_results['model'], feature_columns, "Base Model - Feature Importance",
        os.path.join(charts_dir, "base_model_feature_importance.png")
    )

    # Plot scatter plots
    plot_academic_scatter(
        base_results['true']['train'], base_results['pred']['train'],
        "Base Model - Training Set", base_results['metrics']['train']['r2'],
        os.path.join(charts_dir, "base_model_train_scatter.png")
    )

    plot_academic_scatter(
        base_results['true']['test'], base_results['pred']['test'],
        "Base Model - Test Set", base_results['metrics']['test']['r2'],
        os.path.join(charts_dir, "base_model_test_scatter.png")
    )

    # Plot the side-by-side comparison of the training set and test set
    plot_side_by_side_scatter(
        base_results['true']['train'], base_results['pred']['train'],
        base_results['true']['test'], base_results['pred']['test'],
        "Base Model Evaluation",
        base_results['metrics']['train']['r2'],
        base_results['metrics']['test']['r2'],
        os.path.join(charts_dir, "base_model_side_by_side.png")
    )

    # Export prediction results to CSV
    export_predictions_to_csv(
        base_results['true']['train'], base_results['pred']['train'],
        "base_model_training", charts_dir
    )

    export_predictions_to_csv(
        base_results['true']['test'], base_results['pred']['test'],
        "base_model_test", charts_dir
    )

    # Export feature importance
    export_feature_importance(
        base_results['model'], feature_columns, charts_dir
    )

    # Execute semi-supervised learning with generated data
    print("\n4. Generating data based on the training set and executing semi-supervised learning...")
    try:
        best_ssl_model, generated_file, excel_path, iterations_excel, csv_files = run_with_generated_data(
            base_results, X_train, y_train, X_test, y_test, feature_columns, output_dirs, original_df=df_cleaned
        )

        if best_ssl_model is not None:
            # Save MAE-best SSL model analysis results
            try:
                # Directly test the performance of the best model on the test set
                y_test_pred_ssl = best_ssl_model.predict(X_test)
                test_r2_ssl = r2_score(y_test, y_test_pred_ssl)
                test_mae_ssl = mean_absolute_error(y_test, y_test_pred_ssl)
                test_rmse_ssl = np.sqrt(mean_squared_error(y_test, y_test_pred_ssl))

                print("\nSSL model test set metrics based on in-memory data:")
                print(f"  Test set R²: {test_r2_ssl:.9f}, MAE: {test_mae_ssl:.9f}, RMSE: {test_rmse_ssl:.9f}")

                # Add SSL model metrics to metrics_df
                metrics_df = pd.concat([metrics_df, pd.DataFrame([{
                    "Model": "SSL_RandomForest_Best_MAE",
                    "Training_R2": "N/A",  # Training set metrics are no longer calculated
                    "Training_MAE": "N/A",
                    "Training_RMSE": "N/A",
                    "Validation_R2": "N/A",  # The SSL model has no validation set results
                    "Validation_MAE": "N/A",
                    "Validation_RMSE": "N/A",
                    "Testing_R2": test_r2_ssl,
                    "Testing_MAE": test_mae_ssl,
                    "Testing_RMSE": test_rmse_ssl
                }])], ignore_index=True)
            except Exception as e:
                print(f"An error occurred while calculating SSL model test set metrics: {e}")
                traceback.print_exc()

        # Save complete metrics - 9-decimal precision
        metrics_df.to_csv(result_file, index=False, float_format='%.9f')
        print(f"\nModel performance metrics have been saved to: {result_file}")

        # Read final high-precision metrics from Parquet files
        try:
            print("\nGetting final model precise metrics...")
            # Build the parquet directory path
            parquet_dir = os.path.join(data_dir, 'parquet_data')
            if os.path.exists(parquet_dir):
                # Try to directly read the saved metrics
                metrics_file = os.path.join(parquet_dir, 'metrics.parquet')
                if os.path.exists(metrics_file):
                    parquet_metrics_df = pd.read_parquet(metrics_file)

                    # Get metrics for the final training set (original + pseudo-labels)
                    final_train_metrics = parquet_metrics_df[parquet_metrics_df['dataset'] == 'final_train']
                    final_train_r2 = float(final_train_metrics[final_train_metrics['metric'] == 'r2']['value'])
                    final_train_mae = float(final_train_metrics[final_train_metrics['metric'] == 'mae']['value'])
                    final_train_rmse = float(final_train_metrics[final_train_metrics['metric'] == 'rmse']['value'])

                    # Get metrics for the test set
                    test_metrics = parquet_metrics_df[parquet_metrics_df['dataset'] == 'test']
                    final_test_r2 = float(test_metrics[test_metrics['metric'] == 'r2']['value'])
                    final_test_mae = float(test_metrics[test_metrics['metric'] == 'mae']['value'])
                    final_test_rmse = float(test_metrics[test_metrics['metric'] == 'rmse']['value'])

                    # Read metadata to obtain the number of samples
                    metadata_file = os.path.join(parquet_dir, 'metadata.parquet')
                    if os.path.exists(metadata_file):
                        metadata_df = pd.read_parquet(metadata_file)
                        train_samples = metadata_df['train_samples'].iloc[
                            0] if 'train_samples' in metadata_df.columns else 0
                        pseudo_samples = metadata_df['pseudo_samples'].iloc[
                            0] if 'pseudo_samples' in metadata_df.columns else 0
                        test_samples = metadata_df['test_samples'].iloc[
                            0] if 'test_samples' in metadata_df.columns else 0
                    else:
                        # If there is no metadata, try to obtain it from the data files
                        final_train_df = pd.read_parquet(os.path.join(parquet_dir, 'final_ssl_train.parquet'))
                        test_df = pd.read_parquet(os.path.join(parquet_dir, 'test.parquet'))
                        original_train_df = pd.read_parquet(os.path.join(parquet_dir, 'original_train.parquet'))

                        train_samples = len(original_train_df)
                        test_samples = len(test_df)
                        pseudo_samples = len(final_train_df) - len(original_train_df)

                    # Output high-precision metrics
                    print("\nFinal SSL model high-precision metrics (read from Parquet files):")
                    print(
                        f"  Final training set (original + {pseudo_samples} pseudo-labels) R²: {final_train_r2:.9f}, MAE: {final_train_mae:.9f}, RMSE: {final_train_rmse:.9f}")
                    print(f"  Test set R²: {final_test_r2:.9f}, MAE: {final_test_mae:.9f}, RMSE: {final_test_rmse:.9f}")

                    # Save final metrics to CSV file
                    final_metrics_file = os.path.join(run_dir, "final_precise_metrics.csv")
                    final_metrics_df = pd.DataFrame({
                        "Metric": ["Training R²", "Training MAE", "Training RMSE",
                                   "Testing R²", "Testing MAE", "Testing RMSE"],
                        "Value": [final_train_r2, final_train_mae, final_train_rmse,
                                  final_test_r2, final_test_mae, final_test_rmse],
                        "Dataset": ["Final Training Set (Original + Pseudo)", "Final Training Set (Original + Pseudo)",
                                    "Final Training Set (Original + Pseudo)",
                                    "Test Set", "Test Set", "Test Set"],
                        "Total Samples": [train_samples + pseudo_samples, train_samples + pseudo_samples,
                                          train_samples + pseudo_samples,
                                          test_samples, test_samples, test_samples],
                        "Original Samples": [train_samples, train_samples, train_samples, 0, 0, 0],
                        "Pseudo-label Samples": [pseudo_samples, pseudo_samples, pseudo_samples, 0, 0, 0]
                    })

                    final_metrics_df.to_csv(final_metrics_file, index=False, float_format='%.9f')
                    print(f"\nFinal model precise metrics have been saved to: {final_metrics_file}")

                    # Add high-precision metrics to the performance comparison Excel
                    if os.path.exists(excel_path):
                        try:
                            # Read Excel
                            wb = openpyxl.load_workbook(excel_path)
                            ws = wb.active

                            # Add a new column - high-precision final model
                            max_row = ws.max_row
                            max_col = ws.max_column + 1

                            # Add column header
                            ws.cell(row=1, column=max_col).value = "Final Parquet Model"

                            # Iterate through rows, find corresponding metrics, and fill them
                            for row in range(2, max_row + 1):
                                metric = ws.cell(row=row, column=1).value
                                if metric == "Training R²":
                                    ws.cell(row=row, column=max_col).value = final_train_r2
                                elif metric == "Training MAE":
                                    ws.cell(row=row, column=max_col).value = final_train_mae
                                elif metric == "Training RMSE":
                                    ws.cell(row=row, column=max_col).value = final_train_rmse
                                elif metric == "Testing R²":
                                    ws.cell(row=row, column=max_col).value = final_test_r2
                                elif metric == "Testing MAE":
                                    ws.cell(row=row, column=max_col).value = final_test_mae
                                elif metric == "Testing RMSE":
                                    ws.cell(row=row, column=max_col).value = final_test_rmse
                                else:
                                    ws.cell(row=row, column=max_col).value = "N/A"

                            # Style settings
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='center')

                            # Save Excel
                            wb.save(excel_path)
                            print(f"\nHigh-precision metrics have been added to the performance comparison table: {excel_path}")
                        except Exception as e:
                            print(f"An error occurred while updating the Excel file: {e}")
                            traceback.print_exc()

                    # Create the three-model comparison CSV file
                    try:
                        model_comparison_file = os.path.join(run_dir, "three_models_comparison.csv")
                        three_models_df = pd.DataFrame({
                            "Model": ["Base Model", "SSL Model (MAE Best)", "Final Parquet Model"],
                            "Training_R2": [
                                base_results['metrics']['train']['r2'],
                                "N/A",  # No longer use in-memory training set metrics
                                final_train_r2
                            ],
                            "Training_MAE": [
                                base_results['metrics']['train']['mae'],
                                "N/A",  # No longer use in-memory training set metrics
                                final_train_mae
                            ],
                            "Training_RMSE": [
                                base_results['metrics']['train']['rmse'],
                                "N/A",  # No longer use in-memory training set metrics
                                final_train_rmse
                            ],
                            "Test_R2": [
                                base_results['metrics']['test']['r2'],
                                test_r2_ssl,
                                final_test_r2
                            ],
                            "Test_MAE": [
                                base_results['metrics']['test']['mae'],
                                test_mae_ssl,
                                final_test_mae
                            ],
                            "Test_RMSE": [
                                base_results['metrics']['test']['rmse'],
                                test_rmse_ssl,
                                final_test_rmse
                            ],
                            "Pseudo_Samples": [
                                0,
                                "Unknown",  # No longer have in-memory pseudo-label sample count information
                                pseudo_samples
                            ]
                        })
                        three_models_df.to_csv(model_comparison_file, index=False, float_format='%.9f')
                        print(f"\nThree-model comparison has been saved to: {model_comparison_file}")
                    except Exception as e:
                        print(f"An error occurred while saving the three-model comparison: {e}")
                        traceback.print_exc()
                else:
                    print(f"Metrics file not found: {metrics_file}")
            else:
                print(f"Parquet directory does not exist: {parquet_dir}")
        except Exception as e:
            print(f"An error occurred while getting final precise metrics: {e}")
            traceback.print_exc()

        print("\nTraining completed! All outputs are saved in:", run_dir)

        if generated_file:
            print(f"Generated data file: {os.path.basename(generated_file)}")

        print(f"Performance metrics CSV file: {os.path.basename(result_file)}")

        if excel_path:
            print(f"Performance metrics Excel file: {os.path.basename(excel_path)}")

        if iterations_excel:
            print(f"SSL iteration metrics Excel file: {os.path.basename(iterations_excel)}")

    except Exception as e:
        print(f"An error occurred while running semi-supervised learning: {e}")
        traceback.print_exc()
        # Save base model metrics
        metrics_df.to_csv(result_file, index=False, float_format='%.9f')
        print(f"Base model metrics have been saved to: {result_file}")


if __name__ == "__main__":
    main()
